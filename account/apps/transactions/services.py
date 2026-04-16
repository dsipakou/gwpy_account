import copy
import datetime
import io
from collections import defaultdict
from typing import Literal
from uuid import UUID

import openpyxl
from django.db.models import QuerySet
from openpyxl.utils import get_column_letter

from categories import constants as category_constants
from currencies.models import Currency
from rates.models import Rate
from rates.utils import generate_amount_map
from transactions.entities import (
    GroupedByCategory,
    GroupedByMonth,
    GroupedByParent,
    TransactionAccountDetails,
    TransactionBudgetDetails,
    TransactionCategoryDetails,
    TransactionCurrencyDetails,
    TransactionItem,
    TransactionSpentInCurrencyDetails,
)
from transactions.models import Transaction, TransactionMulticurrency
from workspaces.models import Workspace


class TransactionService:
    @classmethod
    def create_transaction_multicurrency_amount(
        cls, uuids: list[UUID], workspace: Workspace
    ):
        amount_mapping = {}
        transactions = Transaction.objects.select_related("currency").filter(
            uuid__in=uuids, workspace=workspace
        )
        dates = transactions.values_list("transaction_date", flat=True).distinct()
        rates_on_date = Rate.objects.filter(rate_date__in=dates, workspace=workspace)
        for transaction in transactions:
            amount_mapping = generate_amount_map(transaction, rates_on_date, workspace)

            TransactionMulticurrency.objects.update_or_create(
                transaction=transaction, defaults={"amount_map": amount_mapping}
            )

    @classmethod
    def get_transaction(cls, transaction: Transaction) -> Transaction | None:
        category_details = TransactionCategoryDetails(
            name=transaction.category.name,
            parent=(
                transaction.category.parent.uuid
                if transaction.category.type == category_constants.EXPENSE
                else ""
            ),
            parent_name=(
                transaction.category.parent.name
                if transaction.category.type == category_constants.EXPENSE
                else ""
            ),
        )
        account_details = TransactionAccountDetails(
            title=transaction.account.title,
        )
        currency_details = TransactionCurrencyDetails(
            sign=transaction.currency.sign,
        )
        spent_details = (
            TransactionSpentInCurrencyDetails(transaction.multicurrency.amount_map)
            if hasattr(transaction, "multicurrency")
            else None
        )

        budget_details = None
        if transaction.budget:
            budget_details = TransactionBudgetDetails(
                title=transaction.budget.title,
            )

        return TransactionItem(
            uuid=transaction.uuid,
            user=transaction.user.uuid,
            category=transaction.category.uuid,
            category_details=category_details,
            budget=transaction.budget.uuid if transaction.budget else None,
            budget_details=budget_details,
            currency=transaction.currency.uuid,
            currency_details=currency_details,
            amount=transaction.amount,
            spent_in_currencies=spent_details,
            account=transaction.account.uuid,
            account_details=account_details,
            description=transaction.description,
            transaction_date=transaction.transaction_date,
            created_at=transaction.created_at,
            modified_at=transaction.modified_at,
        )

    @classmethod
    def group_by_month(cls, transactions: list[Transaction]) -> list[GroupedByMonth]:
        grouped_by_month = {}
        for transaction in transactions:
            transaction_details: TransactionItem = cls.get_transaction(transaction)
            date = transaction_details["transaction_date"]
            formatted_date = f"{date.year}-{date.month}"
            grouped_by_month[formatted_date] = grouped_by_month.get(formatted_date)
            if not grouped_by_month[formatted_date]:
                grouped_by_month[formatted_date] = {
                    "month": date.month,
                    "year": date.year,
                    "spent_in_base_currency": transaction_details[
                        "spent_in_base_currency"
                    ],
                    "spent_in_currencies": transaction_details["spent_in_currencies"],
                }
            else:
                grouped_by_month[formatted_date]["spent_in_base_currency"] += (
                    transaction_details["spent_in_base_currency"]
                )
                for currency in grouped_by_month[formatted_date]["spent_in_currencies"]:
                    grouped_by_month[formatted_date]["spent_in_currencies"][
                        currency
                    ] += transaction_details["spent_in_currencies"][currency]

    @classmethod
    def group_by_category(cls, transactions: QuerySet) -> GroupedByCategory:
        grouped_by_category = {}
        for transaction in transactions:
            transaction_details: TransactionItem = cls.get_transaction(transaction)
            category_name = transaction_details["category_details"]["name"]
            parent_name = transaction_details["category_details"]["parent_name"]
            if category_name not in grouped_by_category:
                grouped_by_category[category_name] = GroupedByCategory(
                    category_name=category_name,
                    parent_name=parent_name,
                    spent_in_base_currency=transaction_details[
                        "spent_in_base_currency"
                    ],
                    spent_in_currencies=copy.deepcopy(
                        transaction_details["spent_in_currencies"]
                    ),
                    items=[transaction_details],
                )
                continue

            grouped_by_category[category_name]["items"].append(transaction_details)

            grouped_by_category[category_name]["spent_in_base_currency"] += (
                transaction_details["spent_in_base_currency"]
            )

            for currency, value in grouped_by_category[category_name][
                "spent_in_currencies"
            ].items():
                grouped_by_category[category_name]["spent_in_currencies"][currency] += (
                    value
                )
        return grouped_by_category

    @classmethod
    def group_by_parent(
        cls, grouped_by_category: dict[str, GroupedByCategory]
    ) -> GroupedByParent:
        grouped_by_parent = {}
        for _, category in sorted(grouped_by_category.items()):
            parent_name = category["parent_name"]
            if parent_name not in grouped_by_parent:
                grouped_by_parent[parent_name] = GroupedByParent(
                    category_name=parent_name,
                    spent_in_base_currency=category["spent_in_base_currency"],
                    spent_in_currencies=copy.deepcopy(category["spent_in_currencies"]),
                    items=[category],
                )
                continue

            grouped_by_parent[parent_name]["items"].append(category)

            grouped_by_parent[parent_name]["spent_in_base_currency"] += category[
                "spent_in_base_currency"
            ]

            for currency, value in grouped_by_parent[parent_name][
                "spent_in_currencies"
            ].items():
                grouped_by_parent[parent_name]["spent_in_currencies"][currency] += value
        return grouped_by_parent

    @classmethod
    def load_transaction(cls, transaction_uuid: str) -> TransactionItem:
        transaction = Transaction.objects.get(uuid=transaction_uuid)
        return cls.get_transaction(transaction)

    @classmethod
    def load_transactions(
        cls,
        queryset: QuerySet,
        *,
        limit: int = 15,
        order_by: str = "created_at",
        category_type: Literal["INC", "EXP"] = category_constants.EXPENSE,
        date_from: str | None = None,
        date_to: str | None = None,
    ) -> list[TransactionItem]:
        qs = queryset.filter(category__type=category_type).select_related("category")

        if date_from and date_to:
            qs = qs.filter(
                transaction_date__lte=date_to,
                transaction_date__gte=date_from,
            )

        qs = qs.order_by(f"-{order_by}").select_related(
            "multicurrency", "category", "account", "budget"
        )[:limit]

        return cls.proceed_transactions(qs)

    @classmethod
    def proceed_transactions(
        cls,
        queryset: QuerySet,
    ) -> list[TransactionItem]:
        transactions = []
        for transaction in queryset:
            transactions.append(cls.get_transaction(transaction))
        return transactions

    @classmethod
    def build_transactions_excel(
        cls, transactions: list[TransactionItem], workspace: Workspace
    ) -> io.BytesIO:
        base_currency = Currency.objects.get(workspace=workspace, is_base=True)

        # group_data[parent][child][date] = list of amounts
        group_data: dict[str, dict[str, dict[datetime.date, list[float]]]] = (
            defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
        )
        all_dates_set: set[datetime.date] = set()

        for t in transactions:
            spent = t.get("spent_in_currencies") or {}
            amount = spent.get(base_currency.code, t["amount"])
            child = t["category_details"]["name"]
            parent = t["category_details"]["parent_name"] or child
            txn_date = t["transaction_date"]
            group_data[parent][child][txn_date].append(amount)
            all_dates_set.add(txn_date)

        all_dates = sorted(all_dates_set)
        date_col_count = len(all_dates)
        sorted_parents = sorted(group_data.keys())

        # "Другое" sorts last within each group
        def child_sort_key(name: str) -> tuple:
            return (name == "Другое", name)

        # Pre-calculate row numbers so SUM formulas can reference exact ranges
        # Layout: row1=title, row2=grand total, row3=weekdays, row4+=groups
        HEADER_ROWS = 3
        current_row = HEADER_ROWS + 1
        structure: list[tuple] = []
        summary_rows: list[int] = []

        for parent in sorted_parents:
            children = sorted(group_data[parent].keys(), key=child_sort_key)
            summary_row = current_row + 1
            children_start = current_row + 2
            children_end = children_start + len(children) - 1
            summary_rows.append(summary_row)
            structure.append(("parent", parent, None))
            structure.append(("sum", parent, (children_start, children_end)))
            for child in children:
                structure.append(("child", (parent, child), None))
            current_row = children_end + 1

        quarter = (all_dates[0].month - 1) // 3 + 1 if all_dates else 1

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Transactions"

        # Row 1: title + dates
        ws.append([f"Q {quarter}"] + [f"{d.month}/{d.day}/{d.year}" for d in all_dates])

        # Row 2: grand total = SUM of all group Всего rows
        grand_row: list = [""]
        for col_idx, _ in enumerate(all_dates, start=2):
            col = get_column_letter(col_idx)
            refs = ",".join(f"{col}{r}" for r in summary_rows)
            grand_row.append(f"=SUM({refs})" if refs else None)
        ws.append(grand_row)

        # Row 3: weekday names
        ws.append([""] + [d.strftime("%A") for d in all_dates])

        # Group rows
        for kind, name, extra in structure:
            if kind == "parent":
                ws.append([name] + [None] * date_col_count)
            elif kind == "sum":
                children_start, children_end = extra
                overall_row_data: list = ["Overall"]
                for col_idx, _ in enumerate(all_dates, start=2):
                    col = get_column_letter(col_idx)
                    overall_row_data.append(
                        f"=SUM({col}{children_start}:{col}{children_end})"
                    )
                ws.append(overall_row_data)
            else:  # child
                parent, child = name
                row: list = [child]
                for d in all_dates:
                    amounts = group_data[parent][child].get(d)
                    if amounts:
                        row.append("=" + "+".join(str(round(a, 2)) for a in amounts))
                    else:
                        row.append(None)
                ws.append(row)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    @classmethod
    def load_grouped_transactions(
        cls, queryset, *, date_from: str | None = None, date_to: str | None = None
    ) -> list[GroupedByParent]:
        qs = queryset.order_by("-created_at")
        if date_from:
            qs = qs.filter(transaction_date__gte=date_from)
        if date_to:
            qs = qs.filter(transaction_date__lte=date_to)

        grouped_by_category = cls.group_by_category(qs)
        grouped_by_parent = cls.group_by_parent(grouped_by_category)

        transactions = []
        for _, value in sorted(grouped_by_parent.items()):
            transactions.append(value)
        return transactions


class ReportService:
    @classmethod
    def get_year_report(
        cls, date_from: str, date_to: str, currency_code: str, workspace: str
    ):
        return Transaction.grouped_by_month(
            date_from, date_to, currency_code, workspace
        )

    @classmethod
    def get_chart_report(
        cls,
        qs: QuerySet,
        categories_qs: QuerySet,
        date_from: str,
        date_to: str,
        currency_code: str,
        till_day: int | None = None,
    ):
        grouped_transactions = Transaction.grouped_by_month_and_category(
            qs, date_from, date_to, currency_code, till_day
        )

        grouped_map = cls._get_grouped_map(grouped_transactions)

        start_date = date_from
        end_date = date_to

        month_year_list = []
        while start_date < end_date:
            month_year_list.append(start_date.strftime("%Y-%m"))
            start_date = start_date.replace(day=1) + datetime.timedelta(days=32)
            start_date = start_date.replace(day=1)

        categories_list = categories_qs.filter(parent__isnull=True).order_by("name")

        output = []

        for date_item in month_year_list:
            current_item = {}
            current_item["date"] = date_item
            grouped_item = grouped_map.get(date_item, {})
            current_date_category_list = []
            for category in categories_list:
                categories_map = {}
                categories_map["name"] = category.name
                categories_map["value"] = grouped_item.get(category.name, 0)
                categories_map["category_type"] = str(category.type)
                current_date_category_list.append(categories_map)
            current_item["categories"] = current_date_category_list
            output.append(current_item)

        return output

    @classmethod
    def _get_grouped_map(cls, qs):
        output = {}
        for item in qs:
            key = item["year_month"].strftime("%Y-%m")
            output[key] = output.get(key, {})
            if item["category__parent"]:
                output[key][item["category__parent__name"]] = (
                    output[key].get(item["category__parent__name"], 0)
                    + item["parent_sum"]
                )
            else:
                output[key][item["category__name"]] = (
                    output[key].get(item["category__name"], 0) + item["parent_sum"]
                )

        return output
